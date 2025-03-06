import argparse
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------
# 1. 定义颜色（可根据需求修改）
# ---------------------------------------------------------
INNER_BORDER_COLOR = "9BC2E6"  # 内部边框颜色
OUTER_BORDER_COLOR = "5B9BD5"  # 外部边框颜色
HEADER_FILL_COLOR = "DEEBF7"   # 表头填充颜色
FAILED_FILL_COLOR = "E6B8B7"   # 包含 FAILED rc=1 行的填充颜色

# ---------------------------------------------------------
# 2. 读取 CSV 文件函数
# ---------------------------------------------------------
def read_csv(filename):
    data = []
    with open(filename, newline='', encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            data.append(row)
    return data

# ---------------------------------------------------------
# 3. 创建 Excel 并写入数据
# ---------------------------------------------------------
def write_data_to_sheet(ws, data):
    for row in data:
        ws.append(row)

# ---------------------------------------------------------
# 4. 美化处理函数
# ---------------------------------------------------------
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3

def apply_border(ws, min_row, min_col, max_row, max_col):
    thin_side = Side(style="thin", color=INNER_BORDER_COLOR)
    thick_side = Side(style="thick", color=OUTER_BORDER_COLOR)
    for row in range(min_row, max_row+1):
        for col in range(min_col, max_col+1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left = thick_side if col == min_col else thin_side,
                right = thick_side if col == max_col else thin_side,
                top = thick_side if row == min_row else thin_side,
                bottom = thick_side if row == max_row else thin_side
            )

def fill_header(ws):
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def merge_same_column(ws, col, start_row=2):
    """
    将从 start_row 开始，连续相同的单元格纵向合并，并居中显示。
    """
    merge_start = start_row
    prev_val = ws.cell(row=start_row, column=col).value
    for row in range(start_row+1, ws.max_row+2):
        curr_val = ws.cell(row=row, column=col).value if row <= ws.max_row else None
        if curr_val != prev_val:
            if row - merge_start > 1:
                ws.merge_cells(start_row=merge_start, start_column=col, end_row=row-1, end_column=col)
                cell = ws.cell(row=merge_start, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            merge_start = row
            prev_val = curr_val

def set_all_cells_center(ws):
    """
    将工作表所有单元格内容居中
    """
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

def process_failed_rows(ws):
    """
    对于包含"FAILED rc=1"的行，不进行合并，
    而是为该行所有单元格设置背景填充颜色为 FAILED_FILL_COLOR，并使内容居中。
    """
    max_col = ws.max_column
    for r in range(1, ws.max_row+1):
        row_contains_failed = False
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            if cell.value and "FAILED rc=1" in str(cell.value):
                row_contains_failed = True
                break
        if row_contains_failed:
            for c in range(1, max_col+1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor=FAILED_FILL_COLOR)
                cell.alignment = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------
# 5. 主流程
# ---------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="处理CSV数据并转换为Excel")
    parser.add_argument("--results", default="log.csv", help="指定结果CSV文件路径（默认：log.csv）")
    parser.add_argument("--failed", default="failed.csv", help="指定失败CSV文件路径（默认：failed.csv）")
    parser.add_argument("--output", default="report.xlsx", help="指定输出Excel文件路径（默认：report.xlsx）")
    args = parser.parse_args()

    results_csv = args.results
    failed_csv  = args.failed
    output_excel = args.output

    # 读取 CSV 数据
    data_results = read_csv(results_csv)
    data_failed = read_csv(failed_csv)

    # 创建工作簿和 Sheet
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Log Results"
    ws2 = wb.create_sheet(title="Failed Items")

    # 将数据写入 Sheet
    write_data_to_sheet(ws1, data_results)
    write_data_to_sheet(ws2, data_failed)

    # 美化处理每个 Sheet
    for ws in [ws1, ws2]:
        fill_header(ws)                                  # 设置表头背景及加粗
        merge_same_column(ws, 1, start_row=1)              # 合并第一列相邻相同单元格（从第一行开始）
        merge_same_column(ws, 2, start_row=2)              # 合并第二列相邻相同单元格（从第二行开始）
        process_failed_rows(ws)                          # 对包含 "FAILED rc=1" 的行设置填充颜色，不进行合并
        auto_adjust_column_width(ws)                     # 自动调整列宽
        apply_border(ws, 1, 1, ws.max_row, ws.max_column)  # 添加边框
        set_all_cells_center(ws)                         # 所有单元格居中

    # 保存结果
    wb.save(output_excel)
    print(f"已生成 {output_excel}，包含两个 Sheet（Log Results 和 Failed Items）")

if __name__ == "__main__":
    main()

